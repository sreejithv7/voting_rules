import openpyxl
import pprint
# import os

# Go to folder where excel file is located
# os.chdir('/Users/sreejith/Downloads')
# print(os.getcwd())


def generatePreferences(values):

    # Initilizing the dictionary with agents as keys, based on max row in sheet
    # preferences_dict = {key: [] for key in range(1, sheet.max_row+1)}

    preferences_dict = {}

    for row in range(1, values.max_row+1):
        key = row
        if key not in preferences_dict.keys():
            preferences_dict[key] = []

        # Looping through the entire column values for each row
        for col in range(1, values.max_column+1):
            # Appending each column value to the list for each row
            preferences_dict[key].append(values.cell(row, col).value)

    

    # print('\n\nPopulated preferences dictionary: ')
    # pp.pprint(preferences_dict)

    # Create a temporary dictionary to access the list index and values, as 
    temp_dict = {}
    for key, value in preferences_dict.items():
        # temp_dict = dict(enumerate(value, start=1))

        for index, element in enumerate(value):
            # Creating a temp_dict, with index values of list as keys, and list elements as dictionary values
            temp_dict[index+1] = element


        # To get sorted values from dictionary
        # print(sorted(temp_dict.values())[::-1])

        # To get only matching keys of sorted values from dictionary
        # print(sorted(temp_dict, key=temp_dict.get)[::-1])

        # reverse=True, for equal values, gets the key with lowest value first ([::1], sorts that issue)
        # print(sorted(temp_dict, key=temp_dict.get, reverse=True))

        # Updating preferences dictionary with sorted indexes (keys) of temp_dict based on it's values
        preferences_dict[key] = sorted(temp_dict, key=temp_dict.get)[::-1]

        # To get sorted values and matching keys as a tuple from dictionary
        # print(sorted(temp_dict.items(), key=lambda x:x[1])[::-1])

    # print('\n\nSorted preferences dictionary: ')
    # pprint.pprint(preferences_dict)
    return preferences_dict

if __name__ == '__main__':
    wb = openpyxl.load_workbook('/Users/sreejith/Downloads/voting.xlsx')
    sheet = wb['Sheet1']

    # For printing a dictionary properly
    pp = pprint.PrettyPrinter(indent=4, width = 120)

    # pprint.pprint(generatePreferences(sheet))

# pprint.pprint(generatePreferences(sheet))

def dictatorship(preferenceProfile, agent):
     return preferenceProfile[agent][0]

# for agent in range(1, 9):
#     print(f'For agent {agent}: {dictatorship(generatePreferences("/Users/sreejith/Downloads/voting.xlsx"), agent)}')

def plurality(preferences, tiebreak):
    temp_dict = {}
    for key, value in preferences.items():
        temp_dict[key] = value[0]

    print(temp_dict)

    temp_dict2 = {}
    for value in temp_dict.values():
        if value not in temp_dict2.keys():
            temp_dict2[value] = 1
        else:
            temp_dict2[value] += 1

    print(temp_dict2)
    list_of_keys = []
    max_value = max(temp_dict2.values())
    for key, value in temp_dict2.items():
        if value == max_value:
            list_of_keys.append(key)
    print(list_of_keys)
    if tiebreak == 'max':
        print(max(list_of_keys))
    elif tiebreak == 'min':
        print(min(list_of_keys))
    else:
        agent = tiebreak
        temp_dict3 = {}
        for item in list_of_keys:
            temp_dict3[item] = preferences[agent].index(item)
        print(preferences[agent])
        print(temp_dict3)
        print(min(temp_dict3, key=temp_dict3.get))
        # print(max(temp_dict3, key=temp_dict3.get))



    

    # print(max(temp_dict2, key=temp_dict2.get))
    # print(max(temp_dict2, key=lambda key: temp_dict2[key]))




    if tiebreak == 'max':
        print(max())




    # for key, value in preferences.items():
    #     preferences[key] = [0 for x in value if value.index]





# preferences_dictionary = {1: [4, 2, 1, 3], 
#                           2: [4, 3, 1, 2],
#                           3: [4, 3, 1, 2],
#                           4: [1, 3, 4, 2],
#                           5: [2, 3, 4, 1],
#                           6: [2, 1, 3, 4],
#                           7: [4, 2, 3, 1],
#                           8: [4, 2, 1, 3]}

# preferences_dictionary = {1: [4, 2, 1, 3], 
#                           2: [4, 3, 1, 2],
#                           3: [5, 3, 1, 4],
#                           4: [1, 3, 4, 2],
#                           5: [5, 3, 4, 1],
#                           6: [5, 1, 3, 4],
#                           7: [1, 2, 3, 4],
#                           8: [4, 2, 1, 3]}

preferences_dictionary = {1: [4, 2, 1, 3], 
                          2: [4, 3, 1, 2],
                          3: [2, 3, 1, 4],
                          4: [1, 3, 4, 2],
                          5: [2, 3, 4, 1],
                          6: [2, 1, 3, 4],
                          7: [1, 2, 3, 4],
                          8: [4, 2, 1, 3]}

plurality(preferences_dictionary, 'max')
plurality(preferences_dictionary, 'min')
plurality(preferences_dictionary, 1)



